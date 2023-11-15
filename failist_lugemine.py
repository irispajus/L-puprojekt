import openpyxl

exceli_asukoht_PCs = 'siia lisa oma faili nimi, asukoht arvutis'

workbook = openpyxl.load_workbook(exceli_asukoht_PCs)
worksheet = workbook.active

#küsib lihtsalt märksõna mida failist otsida, võib olla nii zanr, osa pealkirjast kui ka artist
kasutaja_sisestatud_laulunimi = input("Sisesta laulu nimi ")

#otsib üles tabelist lahtri kaalupunktidega
kaalupunktide_tulp = 'kaalupunktid' in [cell.value for cell in worksheet[1]]

if kaalupunktide_tulp:
    kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

#käib kõik read läbi, kasutab ka funktsiooni lowercase, et kaotada suur- ja väiketähtede tundlikkus ära
for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
    for cell in row:
        if kasutaja_sisestatud_laulunimi.lower() in str(cell.value).lower():
            # liidab kaalupunktide lahtrisse väärtuse +1, vahet pole kas seal juba on juba üks väärtus või mitte
            kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
            worksheet.cell(row=row_index, column= kaalupunktide_tulp_otsing, value= kaalupunktid + 1)
              
#salvestab tehtud muudatused
workbook.save('siia lisa oma faili nimi, asukoht arvutis')




