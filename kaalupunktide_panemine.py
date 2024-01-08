import openpyxl

exceli_asukoht_PCs = 'C:/Users/irispajus/Desktop/PY/andmebaas.xlsx'

workbook = openpyxl.load_workbook(exceli_asukoht_PCs)
worksheet = workbook.active
#tulpade_arv = worksheet.max_column

kasutaja_sisestatud_laulunimi = input("Sisesta laulu nimi ")
kasutaja_sisestatud_lauluartist = input("Sisesta laulu artist")
kasutaja_sisestatud_lauluzanr = input ("Sisesta laulu zanr")
kasutaja_sisestatud_lauluaaasta = input ("Sisesta laulu ilmumisaasta")


kaalupunktide_tulp = 'kaalupunktid' in [cell.value for cell in worksheet[1]]

if kaalupunktide_tulp:
    kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

# osa, mis otsib laulu nime järgi, kuhu panna kaalupunktid
for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
    for cell in row:
        if kasutaja_sisestatud_laulunimi.lower() in str(cell.value).lower():
            # liidab kaalupunktide lahtrisse väärtuse +1, vahet pole kas seal juba on juba üks väärtus või mitte
            kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
            worksheet.cell(row=row_index, column= kaalupunktide_tulp_otsing, value= kaalupunktid + 1)
              
if kaalupunktide_tulp:
    kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

# osa, mis otsib artisti nime järgi kuhu panna kaalupunktid
for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
    for cell in row:
        if kasutaja_sisestatud_lauluartist.lower() in str(cell.value).lower():
            # liidab kaalupunktide lahtrisse väärtuse +1, vahet pole kas seal juba on juba üks väärtus või mitte
            kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
            worksheet.cell(row=row_index, column= kaalupunktide_tulp_otsing, value= kaalupunktid + 2)

if kaalupunktide_tulp:
    kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

# osa, mis otsib zanri järgi, kuhu lisada kaalupunkte
for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
    for cell in row:
        if kasutaja_sisestatud_lauluzanr.lower() in str(cell.value).lower():
            # liidab kaalupunktide lahtrisse väärtuse +1, vahet pole kas seal juba on juba üks väärtus või mitte
            kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
            worksheet.cell(row=row_index, column= kaalupunktide_tulp_otsing, value= kaalupunktid + 3)

if kaalupunktide_tulp:
    kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

# osa, mis otsib zanri järgi, kuhu lisada kaalupunkte
for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
    for cell in row:
        if kasutaja_sisestatud_lauluaaasta.lower() in str(cell.value).lower():
            # liidab kaalupunktide lahtrisse väärtuse +1, vahet pole kas seal juba on juba üks väärtus või mitte
            kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
            worksheet.cell(row=row_index, column= kaalupunktide_tulp_otsing, value= kaalupunktid + 3)


workbook.save('C:/Users/irispajus/Desktop/PY/andmebaas.xlsx')




