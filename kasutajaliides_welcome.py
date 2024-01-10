import os
import shutil
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import random
import openpyxl

def sort_and_save_excel(excel_file_path):
    df = pd.read_excel(excel_file_path)

    kaalupunktide_sorteerimine = 'kaalupunktid'

    # bubble sortimine
    n = len(df)
    for i in range(n):
        for j in range(0, n-i-1):
            if pd.isna(df.loc[j, kaalupunktide_sorteerimine]) or (pd.notna(df.loc[j+1, kaalupunktide_sorteerimine]) and df.loc[j, kaalupunktide_sorteerimine] < df.loc[j+1, kaalupunktide_sorteerimine]):
                # Swap only the 'kaalupunktid' values
                df.loc[j, kaalupunktide_sorteerimine], df.loc[j+1, kaalupunktide_sorteerimine] = df.loc[j+1, kaalupunktide_sorteerimine], df.loc[j, kaalupunktide_sorteerimine]

    print(df[kaalupunktide_sorteerimine].dtype)

    df.to_excel(excel_file_path, index=False)

class WelcomeWindow:
    def __init__(self, master, kasutajanimi, user_db_dir):
        self.master = master
        self.user_db_dir = user_db_dir
        master.title("Tere tulemast, {}".format(kasutajanimi))

        # Aknasuuruse määramine
        master.geometry("300x200")

        # Create a directory for user databases if it doesn't exist
        user_db_dir = 'C:/Users/eveli/OneDrive - Tallinna Tehnikaülikool/Algoritmid23/Projekt/UserDatabases/'
        if not os.path.exists(user_db_dir):
            os.makedirs(user_db_dir)

        # Create a copy of the original Excel file for the user
        original_db_path = 'C:/Users/eveli/OneDrive - Tallinna Tehnikaülikool/Algoritmid23/Projekt/andmebaas.xlsx'
        user_db_path = os.path.join(user_db_dir, f'{kasutajanimi}_andmebaas.xlsx')
        if not os.path.exists(user_db_path):
            shutil.copy(original_db_path, user_db_path)

        # Call the sorting function for the user's database
        sort_and_save_excel(user_db_path)

        # Load the user's database
        self.df = pd.read_excel(user_db_path)

        # ridade defineerimine
        self.pealkiri = 0
        self.artist = 1

        # mitu lugu (rida) on failis
        self.lugude_arv = self.df.shape[0]

        # Nuppude konteiner
        button_container = tk.Frame(self.master)
        button_container.pack(pady=10)

        # Like, Skip ja Sulge nupud
        tk.Button(button_container, text="Like", command=self.like_action).pack(side=tk.LEFT, padx=10)
        tk.Button(button_container, text="Skipp", command=self.skipp_action).pack(side=tk.LEFT, padx=10)
        tk.Button(button_container, text="Sulge", command=self.master.destroy).pack(side=tk.RIGHT, padx=10)

        # Alusta esimese looga
        self.show_random_song()

    def show_random_song(self):
        # Kontrolli, kas kõik lood on läbi käidud, alusta uuesti
        if self.lugude_arv == 0:
            messagebox.showinfo("Teade", "Andmebaas on tühi.")
            return

        random_index = random.randint(0, self.lugude_arv - 1)
        self.current_song = self.df.iloc[random_index]

        # Eemalda eelmine artist ja laul
        for widget in self.master.winfo_children():
            if isinstance(widget, tk.Label):
                widget.destroy()

        tk.Label(self.master, text=f"Artist: {self.current_song.iloc[self.artist]}").pack(pady=10)
        tk.Label(self.master, text=f"Laul: {self.current_song.iloc[self.pealkiri]}").pack(pady=5)


    def like_action(self):
        
        # Kasuta hetkel kuvatud laulu andmeid
        kasutaja_sisestatud_laulunimi = self.current_song['pealkiri']
        kasutaja_sisestatud_lauluartist = self.current_song['artist']
        kasutaja_sisestatud_lauluzanr = self.current_song['zanr']
        kasutaja_sisestatud_lauluaaasta = str(self.current_song['aasta'])

        # Uuenda Excel faili
        update_excel(self.user_db_dir, kasutaja_sisestatud_laulunimi, kasutaja_sisestatud_lauluartist,
                      kasutaja_sisestatud_lauluzanr, kasutaja_sisestatud_lauluaaasta)

        # Näita uut lugu
        self.show_random_song()

    def skipp_action(self):

        # Näita uut lugu
        self.show_random_song() 


def update_excel(user_db_dir, laulu_nimi, laulu_artist, laulu_zanr, laulu_aasta):
    workbook = openpyxl.load_workbook(user_db_dir)
    worksheet = workbook.active

    kaalupunktide_tulp = 'kaalupunktid' in [cell.value for cell in worksheet[1]]

    if kaalupunktide_tulp:
        kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

    # osa, mis otsib laulu nime järgi, kuhu panna kaalupunktid
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for cell in row:
            if laulu_nimi.lower() in str(cell.value).lower():
                kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
                worksheet.cell(row=row_index, column=kaalupunktide_tulp_otsing, value=kaalupunktid + 1)

    if kaalupunktide_tulp:
        kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

    # osa, mis otsib artisti nime järgi kuhu panna kaalupunktid
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for cell in row:
            if laulu_artist.lower() in str(cell.value).lower():
                kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
                worksheet.cell(row=row_index, column=kaalupunktide_tulp_otsing, value=kaalupunktid + 2)

    if kaalupunktide_tulp:
        kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

    # osa, mis otsib zanri järgi, kuhu lisada kaalupunkte
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for cell in row:
            if laulu_zanr.lower() in str(cell.value).lower():
                kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
                worksheet.cell(row=row_index, column=kaalupunktide_tulp_otsing, value=kaalupunktid + 3)

    if kaalupunktide_tulp:
        kaalupunktide_tulp_otsing = [cell.value for cell in worksheet[1]].index('kaalupunktid') + 1

    # osa, mis otsib aasta järgi, kuhu lisada kaalupunkte
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        for cell in row:
            if laulu_aasta.lower() in str(cell.value).lower():
                kaalupunktid = row[kaalupunktide_tulp_otsing - 1].value or 0
                worksheet.cell(row=row_index, column=kaalupunktide_tulp_otsing, value=kaalupunktid + 3)

    workbook.save(user_db_dir)
